"""
Excel Link (Master/Child aggregation) — sync child Excel rows into master.
Enterprise: immutable child rows after first sync; strict schema validation; schema hash; audit on reject.
"""
import os
import re
import json
import uuid
import hashlib
import logging
import threading
import unicodedata
from datetime import datetime
from modules.db import get_db

SYSTEM_COLUMNS = ["_row_uuid", "_source_child_file_id", "_source_user_id", "_last_synced_at", "_row_locked_at"]
# Hidden sheet for link metadata so Sheet1 shows only business columns
META_SHEET = "__link_meta"

try:
    from config import VERSIONS_DIR
except ImportError:
    VERSIONS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "versions")

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    load_workbook = None
    get_column_letter = None
    Worksheet = None

# Per-master lock for concurrency (in-process; file lock available via _acquire_file_lock for multi-worker)
_MASTER_LOCKS = {}
_LOCK_GUARD = threading.Lock()


def _get_master_lock(master_file_id):
    with _LOCK_GUARD:
        if master_file_id not in _MASTER_LOCKS:
            _MASTER_LOCKS[master_file_id] = threading.Lock()
        return _MASTER_LOCKS[master_file_id]


def _acquire_file_lock(master_file_id):
    """Acquire cross-process file lock for master sync. Returns (fd, path) or (None, None)."""
    os.makedirs(VERSIONS_DIR, exist_ok=True)
    lock_path = os.path.join(VERSIONS_DIR, f".sync_lock_{master_file_id.replace('/', '_')}")
    try:
        fd = os.open(lock_path, os.O_CREAT | os.O_RDWR)
        import sys
        if sys.platform == "win32":
            try:
                import msvcrt
                msvcrt.locking(fd, msvcrt.LK_LOCK, 1)
            except (ImportError, OSError):
                pass
        else:
            try:
                import fcntl
                fcntl.flock(fd, fcntl.LOCK_EX)
            except (ImportError, OSError):
                pass
        return fd, lock_path
    except (OSError, IOError) as e:
        logging.warning("excel_links file lock acquire failed: %s", e)
        return None, None


def _release_file_lock(fd, lock_path):
    if fd is None:
        return
    try:
        import sys
        if sys.platform == "win32":
            try:
                import msvcrt
                msvcrt.locking(fd, msvcrt.LK_UNLCK, 1)
            except (ImportError, OSError):
                pass
        else:
            try:
                import fcntl
                fcntl.flock(fd, fcntl.LOCK_UN)
            except (ImportError, OSError):
                pass
        os.close(fd)
    except (OSError, IOError):
        pass


def get_link_by_child(child_file_id):
    """Return file_links row if child_file_id is a linked child."""
    db = get_db()
    row = db.execute("""
        SELECT fl.*, s.sheet_name, s.columns_json,
               s.row_uuid_column_name, s.source_child_column_name, s.source_user_column_name,
               s.last_synced_at_column_name, s.row_locked_at_column_name, s.sync_mode, s.schema_hash,
               s.header_row_index
        FROM file_links fl
        LEFT JOIN excel_link_schema s ON s.master_file_id = fl.master_file_id
        WHERE fl.child_file_id = ? AND fl.link_status = 'active'
    """, (child_file_id,)).fetchone()
    db.close()
    return dict(row) if row else None


def is_master_file(master_file_id):
    """True if this file is a master with at least one active child (read-only for users; only sync writes)."""
    db = get_db()
    row = db.execute(
        "SELECT 1 FROM file_links WHERE master_file_id = ? AND link_status = 'active' LIMIT 1",
        (master_file_id,)
    ).fetchone()
    db.close()
    return row is not None


def get_master_path(master_file_id):
    db = get_db()
    row = db.execute("SELECT path FROM files WHERE file_id = ?", (master_file_id,)).fetchone()
    db.close()
    return row["path"] if row and row["path"] and os.path.exists(row["path"]) else None


def get_child_path(child_file_id):
    db = get_db()
    row = db.execute("SELECT path, owner FROM files WHERE file_id = ?", (child_file_id,)).fetchone()
    db.close()
    if not row or not row["path"] or not os.path.exists(row["path"]):
        return None, None
    return row["path"], row["owner"]


def _get_schema_columns(schema_json):
    if not schema_json:
        return []
    try:
        cols = json.loads(schema_json)
        return cols if isinstance(cols, list) else []
    except Exception:
        return []


def get_file_sheets(file_id):
    """Return list of sheet names in the Excel file."""
    path = get_master_path(file_id) or get_child_path(file_id)[0]
    if not path or not load_workbook:
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


def get_master_columns(master_file_id, sheet_name="Sheet1"):
    """Return list of business column names from master's header row (system columns in __link_meta are excluded)."""
    path = get_master_path(master_file_id)
    if not path or not load_workbook:
        return []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return []
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1] if cell.value is not None]
        wb.close()
        business = [str(h).strip() for h in headers if str(h).strip() not in SYSTEM_COLUMNS]
        return business if business else [str(h).strip() for h in headers]
    except Exception:
        return []


def _normalize_header(val):
    """Trim, collapse whitespace, remove invisible chars. Keep Arabic/English."""
    if val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r"\s+", " ", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Cf" and c not in "\u200b\u200c\u200d\ufeff")
    return s.strip()


def _is_valid_header_value(s, min_len=1):
    """Reject purely numeric junk; require some text."""
    if not s or len(s) < min_len:
        return False
    if re.match(r"^[\d\s\.,\-]+$", s):
        return False
    return True


def _extract_row_as_headers(ws, row_idx, max_columns=200):
    """Scan across columns 1..max_columns; collect (col_idx, normalized_value) for non-empty. Preserve order."""
    result = []
    for col in range(1, max_columns + 1):
        try:
            cell = ws.cell(row=row_idx, column=col)
            val = cell.value
        except Exception:
            break
        norm = _normalize_header(val)
        if norm:
            result.append((col, norm))
    return result


def detect_master_headers(file_path, sheet_name="Sheet1", header_row_mode="auto", max_scan_rows=10, max_columns=200, min_business_cols=2):
    """
    Robust header detection. header_row_mode: "row1"|"row2"|"row3"|"auto".
    Auto: scan rows 1..max_scan_rows, find first valid row (>=min_business_cols non-empty, no duplicates, not system-only).
    Returns dict: sheet_name, header_row_index, business_columns, system_columns, canonical_columns, preview_rows, warning.
    Uses read_only=False because openpyxl read_only mode does not support ws.cell() random access.
    """
    if not file_path or not load_workbook:
        return {"ok": False, "warning": "openpyxl not available", "business_columns": [], "canonical_columns": []}
    try:
        wb = load_workbook(file_path, read_only=False, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {"ok": False, "warning": f"Sheet '{sheet_name}' not found", "business_columns": [], "canonical_columns": []}
        ws = wb[sheet_name]
        header_row_idx = None
        raw_headers = []

        if header_row_mode == "row1":
            header_row_idx = 1
            raw_headers = _extract_row_as_headers(ws, 1, max_columns)
        elif header_row_mode == "row2":
            header_row_idx = 2
            raw_headers = _extract_row_as_headers(ws, 2, max_columns)
        elif header_row_mode == "row3":
            header_row_idx = 3
            raw_headers = _extract_row_as_headers(ws, 3, max_columns)
        else:
            for r in range(1, max_scan_rows + 1):
                row_vals = _extract_row_as_headers(ws, r, max_columns)
                if len(row_vals) < min_business_cols:
                    continue
                vals = [v for _, v in row_vals]
                seen = set()
                dup = False
                for v in vals:
                    if v.lower() in seen:
                        dup = True
                        break
                    seen.add(v.lower())
                if dup:
                    continue
                sys_only = all(v in SYSTEM_COLUMNS for v in vals)
                if sys_only:
                    continue
                header_row_idx = r
                raw_headers = row_vals
                break

        if not header_row_idx or not raw_headers:
            wb.close()
            return {
                "ok": False,
                "warning": "No headers found. Put headers in Row 1 (or Row 2/3) or choose the correct sheet.",
                "business_columns": [],
                "system_columns": SYSTEM_COLUMNS,
                "canonical_columns": SYSTEM_COLUMNS,
                "preview_rows": [],
                "header_row_index": 1,
            }

        headers_ordered = [v for _, v in raw_headers]
        seen_lower = {}
        for i, h in enumerate(headers_ordered):
            key = h.lower()
            if key in seen_lower:
                wb.close()
                return {
                    "ok": False,
                    "warning": f"Duplicate column name: {h}",
                    "business_columns": [],
                    "system_columns": SYSTEM_COLUMNS,
                    "canonical_columns": [],
                    "preview_rows": [],
                    "header_row_index": header_row_idx,
                }
            seen_lower[key] = i

        business_columns = [h for h in headers_ordered if h not in SYSTEM_COLUMNS]
        if not business_columns:
            wb.close()
            return {
                "ok": False,
                "warning": "Only system columns found. Add business columns (e.g. first_name, last_name).",
                "business_columns": [],
                "system_columns": SYSTEM_COLUMNS,
                "canonical_columns": [],
                "preview_rows": [],
                "header_row_index": header_row_idx,
            }

        # Canonical = business only; system columns live in __link_meta, not in Sheet1
        canonical_columns = list(business_columns)

        preview_rows = []
        col_by_name = {h: col_idx for col_idx, h in raw_headers}
        for r in range(header_row_idx + 1, min(header_row_idx + 4, ws.max_row + 1)):
            row_dict = {}
            for h in headers_ordered:
                col_idx = col_by_name.get(h)
                if col_idx is None:
                    continue
                try:
                    val = ws.cell(row=r, column=col_idx).value
                    row_dict[h] = val if val is None else str(val)[:50]
                except Exception:
                    pass
            if row_dict:
                preview_rows.append(row_dict)

        wb.close()
        return {
            "ok": True,
            "sheet_name": sheet_name,
            "header_row_index": header_row_idx,
            "business_columns": business_columns,
            "system_columns": SYSTEM_COLUMNS,
            "canonical_columns": canonical_columns,
            "preview_rows": preview_rows,
            "warning": None,
        }
    except Exception as e:
        logging.exception("detect_master_headers: %s", e)
        return {
            "ok": False,
            "warning": str(e),
            "business_columns": [],
            "system_columns": SYSTEM_COLUMNS,
            "canonical_columns": [],
            "preview_rows": [],
        }


def _master_header_row_and_system_cols(master_path, sheet_name, sys_cols):
    """Return (list of header values for row 1, column count) from master."""
    if not master_path or not load_workbook:
        return [], 0
    try:
        wb = load_workbook(master_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return [], 0
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        wb.close()
        # Ensure system columns exist in header list
        seen = {str(h).strip() for h in headers if h is not None}
        for c in sys_cols:
            if c not in seen:
                headers.append(c)
        return headers, len(headers)
    except Exception:
        return [], 0


def propagate_master_schema_to_child(master_file_id, child_file_id):
    """
    Overwrite child's header row (row 1) with master's canonical schema.
    Uses detect_master_headers if header_row_index in schema; else legacy.
    """
    if not load_workbook:
        return {"ok": False, "error": "openpyxl not installed"}
    link = get_link_by_child(child_file_id)
    if not link or link["master_file_id"] != master_file_id:
        return {"ok": False, "error": "Link not found or not this master"}
    master_path = get_master_path(master_file_id)
    child_path, _ = get_child_path(child_file_id)
    if not master_path or not child_path:
        return {"ok": False, "error": "Master or child file not found"}
    sheet_name = link.get("sheet_name") or "Sheet1"
    hdr_idx = link.get("header_row_index") or 1
    header_mode = f"row{hdr_idx}" if hdr_idx in (1, 2, 3) else "auto"
    detected = detect_master_headers(master_path, sheet_name, header_mode)
    if detected.get("ok") and detected.get("canonical_columns"):
        master_headers = detected["canonical_columns"]
    else:
        row_uuid_col = link.get("row_uuid_column_name") or "_row_uuid"
        source_child_col = link.get("source_child_column_name") or "_source_child_file_id"
        source_user_col = link.get("source_user_column_name") or "_source_user_id"
        last_synced_col = link.get("last_synced_at_column_name") or "_last_synced_at"
        row_locked_at_col = link.get("row_locked_at_column_name") or "_row_locked_at"
        sys_cols = [row_uuid_col, source_child_col, source_user_col, last_synced_col, row_locked_at_col]
        master_headers, _ = _master_header_row_and_system_cols(master_path, sheet_name, sys_cols)
    if not master_headers:
        return {"ok": False, "error": "Could not read master headers"}
    # Master headers here = business only (from detect) or legacy full; we write only business to Sheet1
    business_only = [h for h in master_headers if h not in SYSTEM_COLUMNS]
    if not business_only:
        business_only = master_headers
    try:
        wb_child = load_workbook(child_path, read_only=False, data_only=False)
        if sheet_name not in wb_child.sheetnames:
            wb_child.create_sheet(sheet_name)
        ws_child = wb_child[sheet_name]
        # Sheet1: business columns only (row 1)
        for col_idx, val in enumerate(business_only, start=1):
            ws_child.cell(row=1, column=col_idx, value=val)
        # Remove extra columns (e.g. old system columns) so Sheet1 has only business columns
        max_col = ws_child.max_column or len(business_only)
        if max_col > len(business_only):
            ws_child.delete_cols(len(business_only) + 1, max_col - len(business_only))
        # __link_meta: system columns; ensure same number of data rows as Sheet1
        n_data = max(0, (ws_child.max_row or 1) - 1)
        sys_cols = [link.get("row_uuid_column_name") or "_row_uuid", link.get("source_child_column_name") or "_source_child_file_id",
                    link.get("source_user_column_name") or "_source_user_id", link.get("last_synced_at_column_name") or "_last_synced_at",
                    link.get("row_locked_at_column_name") or "_row_locked_at"]
        ws_meta = _get_or_create_meta_sheet(wb_child, sys_cols)
        _ensure_meta_sheet_row_count(ws_meta, sys_cols, n_data, child_file_id, link.get("child_owner_email") or "")
        wb_child.save(child_path)
        wb_child.close()
        return {"ok": True}
    except Exception as e:
        logging.exception("excel_links propagate_master_schema_to_child: %s", e)
        return {"ok": False, "error": str(e)}


def update_schema_from_master(master_file_id, sheet_name="Sheet1", header_row_mode="auto", header_row_index=None):
    """Update excel_link_schema from detect_master_headers. Uses business columns only; system cols appended."""
    path = get_master_path(master_file_id)
    if not path or not load_workbook:
        return False
    try:
        detected = detect_master_headers(path, sheet_name, header_row_mode or "auto")
        if not detected.get("ok") or not detected.get("business_columns"):
            return False
        business_cols = detected["business_columns"]
        hdr_idx = detected.get("header_row_index", header_row_index or 1)
        schema_hash = _compute_schema_hash(business_cols)
        db = get_db()
        db.execute(
            "UPDATE excel_link_schema SET columns_json=?, schema_hash=?, header_row_index=?, updated_at=? WHERE master_file_id=?",
            (json.dumps(business_cols), schema_hash, hdr_idx, datetime.now().strftime("%Y-%m-%d %H:%M"), master_file_id)
        )
        db.commit()
        db.close()
        return True
    except Exception:
        return False


def propagate_master_schema_to_all_children(master_file_id):
    """Push master schema to all active child files. Call after master schema change."""
    schema = get_schema(master_file_id)
    sheet_name = (schema.get("sheet_name") or "Sheet1") if schema else "Sheet1"
    hdr_idx = schema.get("header_row_index") if schema else 1
    header_mode = f"row{hdr_idx}" if hdr_idx in (1, 2, 3) else "auto"
    update_schema_from_master(master_file_id, sheet_name, header_row_mode=header_mode)
    links = list_links(master_file_id)
    results = []
    for fl in links:
        child_id = fl.get("child_file_id")
        if not child_id:
            continue
        r = propagate_master_schema_to_child(master_file_id, child_id)
        results.append({"child_file_id": child_id, "ok": r.get("ok"), "error": r.get("error")})
    return {"ok": True, "results": results}


def _compute_schema_hash(canonical_columns_list):
    """SHA256 of canonical schema array for governance (exact order)."""
    payload = json.dumps(canonical_columns_list, sort_keys=False)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _validate_strict_schema(child_headers, master_schema_ordered, required_sys_cols):
    """
    Strict schema validation: same count, same order, same names, no extras.
    required_sys_cols must all be present (e.g. _row_uuid, _source_child_file_id, _source_user_id, _last_synced_at, _row_locked_at).
    Returns (ok, error_message, schema_diff_summary).
    """
    diff_summary = []
    if len(child_headers) > len(master_schema_ordered):
        diff_summary.append(f"child_columns({len(child_headers)}) > master({len(master_schema_ordered)})")
        return False, "Schema mismatch: child has extra columns. Same count, order and names required.", "; ".join(diff_summary)
    if len(child_headers) != len(master_schema_ordered):
        diff_summary.append(f"count child={len(child_headers)} master={len(master_schema_ordered)}")
        return False, "Schema mismatch: column count or order differs from master.", "; ".join(diff_summary)
    for i, (ch, ms) in enumerate(zip(child_headers, master_schema_ordered)):
        if (ch or "").strip() != (ms or "").strip():
            diff_summary.append(f"pos{i+1}: child={ch!r} master={ms!r}")
            return False, "Schema mismatch: column order or names differ from master.", "; ".join(diff_summary)
    missing_sys = [c for c in required_sys_cols if c not in child_headers]
    if missing_sys:
        diff_summary.append(f"missing_system_columns={missing_sys}")
        return False, f"Missing required system columns: {missing_sys}", "; ".join(diff_summary)
    return True, None, None


def _validate_child_schema(child_headers, schema_columns, sys_cols):
    """Ensure child has required data columns. System columns may be added if missing. (Legacy; strict path uses _validate_strict_schema.)"""
    required = set(schema_columns)
    child_set = set(child_headers)
    missing_data = required - child_set
    if missing_data:
        return False, f"Missing required columns: {sorted(missing_data)}"
    return True, None


def _get_synced_row_locks(child_file_id):
    """Return dict row_uuid -> row_content_hash for rows already locked (immutable)."""
    db = get_db()
    rows = db.execute(
        "SELECT row_uuid, row_content_hash FROM synced_row_locks WHERE child_file_id = ?",
        (child_file_id,)
    ).fetchall()
    db.close()
    return {str(r["row_uuid"]).strip(): r["row_content_hash"] for r in rows} if rows else {}


def _row_content_hash(row_values, headers_ordered, exclude_col_name):
    """Hash row content for immutability check. Excludes exclude_col_name (e.g. _row_locked_at)."""
    parts = []
    for i, h in enumerate(headers_ordered):
        if h == exclude_col_name:
            continue
        v = row_values[i] if i < len(row_values) else None
        parts.append(json.dumps(v, sort_keys=False, default=str))
    payload = "|".join(parts)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _log_sync_rejection(master_file_id, child_file_id, user_id, reason, schema_diff_summary=None):
    """Mandatory audit logging on any sync rejection."""
    try:
        from modules.files import log_audit
        log_audit(
            user_id or "system",
            "excel_link_sync_rejected",
            "file",
            child_file_id,
            details={
                "master_file_id": master_file_id,
                "reason": reason,
                "schema_diff_summary": schema_diff_summary or "",
            },
        )
    except Exception as e:
        logging.warning("excel_links _log_sync_rejection failed: %s", e)


def _ensure_system_columns(ws, header_row_idx, row_uuid_col, source_child_col, source_user_col, last_synced_col, row_locked_at_col, child_file_id, child_owner):
    """Add system columns if missing (including _row_locked_at). Populate row_uuid/source/user; do not set _row_locked_at here."""
    sys_names = [row_uuid_col, source_child_col, source_user_col]
    if last_synced_col:
        sys_names.append(last_synced_col)
    if row_locked_at_col:
        sys_names.append(row_locked_at_col)
    headers = [cell.value for cell in ws[header_row_idx]]
    col_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            col_map[str(h).strip()] = i + 1

    max_col = len(headers)
    for name in sys_names:
        if name not in col_map:
            max_col += 1
            ws.cell(row=header_row_idx, column=max_col, value=name)
            col_map[name] = max_col

    # Populate system column values for each data row (_row_locked_at left empty until first sync)
    for r in range(header_row_idx + 1, ws.max_row + 1):
        ru = ws.cell(row=r, column=col_map.get(row_uuid_col, 0)).value if row_uuid_col in col_map else None
        if not ru or str(ru).strip() == "":
            ws.cell(row=r, column=col_map[row_uuid_col], value=str(uuid.uuid4()))
        sc = ws.cell(row=r, column=col_map.get(source_child_col, 0)).value if source_child_col in col_map else None
        if not sc or str(sc).strip() == "":
            ws.cell(row=r, column=col_map[source_child_col], value=child_file_id)
        su = ws.cell(row=r, column=col_map.get(source_user_col, 0)).value if source_user_col in col_map else None
        if not su and child_owner:
            ws.cell(row=r, column=col_map[source_user_col], value=child_owner)

    return col_map


def _row_to_dict(ws, row_idx, col_map, schema_columns, sys_cols):
    d = {}
    for col_name, col_idx in col_map.items():
        if col_idx:
            val = ws.cell(row=row_idx, column=col_idx).value
            d[col_name] = val
    return d


def _get_or_create_meta_sheet(wb, sys_cols):
    """Get or create __link_meta sheet with system column headers. Returns worksheet."""
    if META_SHEET in wb.sheetnames:
        ws = wb[META_SHEET]
        for c, name in enumerate(sys_cols, start=1):
            if ws.cell(row=1, column=c).value != name:
                ws.cell(row=1, column=c, value=name)
        return ws
    ws = wb.create_sheet(META_SHEET)
    for c, name in enumerate(sys_cols, start=1):
        ws.cell(row=1, column=c, value=name)
    try:
        ws.sheet_state = "hidden"
    except Exception:
        pass
    return ws


def _read_meta_rows(ws_meta, sys_cols, n_rows):
    """Read up to n_rows data rows from meta sheet. Row 1 = headers. Returns list of dicts (one per data row)."""
    out = []
    for r in range(2, min(2 + n_rows, (ws_meta.max_row or 1) + 1)):
        row_dict = {}
        for i, col in enumerate(sys_cols, start=1):
            row_dict[col] = ws_meta.cell(row=r, column=i).value
        out.append(row_dict)
    return out


def _ensure_meta_sheet_row_count(ws_meta, sys_cols, n_data_rows, child_file_id, child_owner):
    """Ensure __link_meta has exactly n_data_rows data rows; fill new/empty rows with uuid, source, user."""
    current_max = ws_meta.max_row or 1
    current_data_rows = max(0, current_max - 1)
    for _ in range(n_data_rows - current_data_rows):
        next_row = current_max + 1
        current_max = next_row
        for i, col in enumerate(sys_cols, start=1):
            if col == "_row_uuid":
                ws_meta.cell(row=next_row, column=i, value=str(uuid.uuid4()))
            elif col == "_source_child_file_id":
                ws_meta.cell(row=next_row, column=i, value=child_file_id or "")
            elif col == "_source_user_id":
                ws_meta.cell(row=next_row, column=i, value=child_owner or "")
            else:
                ws_meta.cell(row=next_row, column=i, value="")
    for r in range(2, 2 + n_data_rows):
        for i, col in enumerate(sys_cols, start=1):
            val = ws_meta.cell(row=r, column=i).value
            if col == "_row_uuid" and (not val or not str(val).strip()):
                ws_meta.cell(row=r, column=i, value=str(uuid.uuid4()))
            elif col == "_source_child_file_id" and (not val or not str(val).strip()) and child_file_id:
                ws_meta.cell(row=r, column=i, value=child_file_id)
            elif col == "_source_user_id" and (not val or not str(val).strip()) and child_owner:
                ws_meta.cell(row=r, column=i, value=child_owner)


def _write_meta_row(ws_meta, row_idx_1based, meta_dict, sys_cols):
    """Write one data row to __link_meta. row_idx_1based is 2-based (row 2 = first data row)."""
    for i, col in enumerate(sys_cols, start=1):
        ws_meta.cell(row=row_idx_1based, column=i, value=meta_dict.get(col))


def sync_child_to_master(child_file_id, triggered_by="callback"):
    """
    Sync child Excel rows into master. Called from OnlyOffice callback or manual Sync Now.
    - Validates schema
    - Ensures _row_uuid, _source_child_file_id, _source_user_id on child rows
    - Upserts into master by _row_uuid
    """
    if not load_workbook:
        logging.warning("excel_links: openpyxl not available")
        return {"ok": False, "error": "openpyxl not installed", "version_created": False}

    link = get_link_by_child(child_file_id)
    if not link:
        return {"ok": True, "skipped": True, "reason": "not a linked child", "version_created": False}

    master_file_id = link["master_file_id"]
    master_path = get_master_path(master_file_id)
    child_path, child_owner = get_child_path(child_file_id)

    if not master_path or not child_path:
        _log_sync(master_file_id, child_file_id, "error", 0, 0, 0, "File path not found", None)
        return {"ok": False, "error": "Master or child file not found", "version_created": False}

    lock = _get_master_lock(master_file_id)
    fd, lock_path = _acquire_file_lock(master_file_id)
    try:
        with lock:
            try:
                sheet_name = link.get("sheet_name") or "Sheet1"
                schema_cols = _get_schema_columns(link.get("columns_json"))
                sync_mode = (link.get("sync_mode") or "append").strip().lower()
                if sync_mode not in ("append", "upsert"):
                    sync_mode = "append"
                row_uuid_col = link.get("row_uuid_column_name") or "_row_uuid"
                source_child_col = link.get("source_child_column_name") or "_source_child_file_id"
                source_user_col = link.get("source_user_column_name") or "_source_user_id"
                last_synced_col = link.get("last_synced_at_column_name") or "_last_synced_at"
                row_locked_at_col = link.get("row_locked_at_column_name") or "_row_locked_at"
                sys_cols = [row_uuid_col, source_child_col, source_user_col, last_synced_col, row_locked_at_col]
                header_row_idx = 1

                # Open child, ensure __link_meta exists; migrate: if Sheet1 has system cols, copy to __link_meta then remove from Sheet1
                wb_child = load_workbook(child_path, read_only=False, data_only=False)
                if sheet_name not in wb_child.sheetnames:
                    wb_child.close()
                    _log_sync(master_file_id, child_file_id, "error", 0, 0, 0, f"Sheet '{sheet_name}' not found in child", None)
                    return {"ok": False, "error": f"Sheet '{sheet_name}' not found", "version_created": False}
                ws_child = wb_child[sheet_name]
                max_col = ws_child.max_column or 200
                child_headers_raw = []
                col_map_sheet = {}
                for c in range(1, max_col + 1):
                    val = ws_child.cell(row=header_row_idx, column=c).value
                    if val is not None and str(val).strip():
                        h = str(val).strip()
                        child_headers_raw.append(h)
                        col_map_sheet[h] = c
                sys_in_sheet = [h for h in child_headers_raw if h in SYSTEM_COLUMNS]
                business_in_sheet = [h for h in child_headers_raw if h not in SYSTEM_COLUMNS]
                n_data_rows = max(0, (ws_child.max_row or 1) - 1)
                ws_meta = _get_or_create_meta_sheet(wb_child, sys_cols)
                if sys_in_sheet and n_data_rows > 0:
                    # Migrate: copy system column values from Sheet1 to __link_meta
                    for r in range(2, 2 + n_data_rows):
                        for si, sc in enumerate(sys_cols, start=1):
                            cidx = col_map_sheet.get(sc)
                            if cidx is not None:
                                ws_meta.cell(row=r, column=si, value=ws_child.cell(row=r, column=cidx).value)
                _ensure_meta_sheet_row_count(ws_meta, sys_cols, n_data_rows, child_file_id, child_owner or "")
                if sys_in_sheet:
                    # Remove system columns from Sheet1 so user only sees business columns
                    cols_to_del = sorted([col_map_sheet[h] for h in SYSTEM_COLUMNS if h in col_map_sheet], reverse=True)
                    for c in cols_to_del:
                        ws_child.delete_cols(c, 1)
                wb_child.save(child_path)
                wb_child.close()

                # Reload child: Sheet1 = business only; read data and meta
                wb_child = load_workbook(child_path, read_only=True, data_only=True)
                ws_child = wb_child[sheet_name]
                child_headers = [cell.value for cell in ws_child[header_row_idx] if cell.value is not None]
                child_headers = [str(h).strip() for h in child_headers]

                # --- Schema: business columns only in Sheet1; validate and hash ---
                stored_schema_hash = link.get("schema_hash")
                if stored_schema_hash:
                    current_schema_hash = _compute_schema_hash(schema_cols)
                    if current_schema_hash != stored_schema_hash:
                        wb_child.close()
                        _log_sync_rejection(master_file_id, child_file_id, child_owner, "Schema hash mismatch", f"stored={stored_schema_hash[:16]}... current={current_schema_hash[:16]}...")
                        _log_sync(master_file_id, child_file_id, "error", 0, 0, 0, "Schema hash mismatch", None)
                        return {"ok": False, "error": "Schema hash mismatch. Run 'Push schema to all children' from master.", "version_created": False}

                ok, err, diff_summary = _validate_strict_schema(child_headers, schema_cols, [])
                if not ok:
                    wb_child.close()
                    _log_sync_rejection(master_file_id, child_file_id, child_owner, err, diff_summary)
                    _log_sync(master_file_id, child_file_id, "error", 0, 0, 0, err, None)
                    return {"ok": False, "error": err, "version_created": False}

                data_rows = list(ws_child.iter_rows(min_row=header_row_idx + 1, values_only=True))
                headers_ordered = [str(h).strip() if h is not None else "" for h in child_headers]
                col_idx = {h: i for i, h in enumerate(headers_ordered) if h}
                n_data = len(data_rows)
                wb_child.close()

                # Read meta rows from __link_meta (by row index)
                wb_child2 = load_workbook(child_path, read_only=True, data_only=True)
                ws_meta_r = wb_child2[META_SHEET] if META_SHEET in wb_child2.sheetnames else None
                meta_rows = _read_meta_rows(ws_meta_r, sys_cols, n_data) if ws_meta_r else []
                while len(meta_rows) < n_data:
                    meta_rows.append({})
                wb_child2.close()

                def get_val(row, name):
                    idx = col_idx.get(name)
                    if idx is not None and idx < len(row):
                        return row[idx]
                    return None

                # --- Immutable rows: use meta for row_uuid and row_locked_at ---
                row_locks = _get_synced_row_locks(child_file_id)
                for row_idx, row in enumerate(data_rows):
                    if not row or all(c is None or str(c).strip() == "" for c in row):
                        continue
                    meta = meta_rows[row_idx] if row_idx < len(meta_rows) else {}
                    ru = (meta.get(row_uuid_col) or get_val(row, row_uuid_col) or "")
                    ru = str(ru).strip() if ru else ""
                    if not ru:
                        continue
                    locked_at_val = meta.get(row_locked_at_col) or get_val(row, row_locked_at_col)
                    if locked_at_val is not None and str(locked_at_val).strip():
                        if ru in row_locks:
                            current_hash = _row_content_hash(row, headers_ordered, row_locked_at_col)
                            if row_locks[ru] != current_hash:
                                _log_sync_rejection(master_file_id, child_file_id, child_owner, "Row is immutable after first sync.", f"row_uuid={ru}")
                                _log_sync(master_file_id, child_file_id, "error", 0, 0, 0, "Row is immutable after first sync.", None)
                                return {"ok": False, "error": "Row is immutable after first sync.", "version_created": False}

                # Load master: Sheet1 = business only; __link_meta = system
                wb_master = load_workbook(master_path, read_only=False, data_only=False)
                if sheet_name not in wb_master.sheetnames:
                    wb_master.create_sheet(sheet_name)
                ws_master = wb_master[sheet_name]
                m_headers = []
                col_map_m = {}
                for c in range(1, (ws_master.max_column or 200) + 1):
                    val = ws_master.cell(row=1, column=c).value
                    if val is not None and str(val).strip():
                        h = str(val).strip()
                        m_headers.append(h)
                        col_map_m[h] = c
                m_business = [h for h in m_headers if h not in SYSTEM_COLUMNS]
                if not m_business:
                    m_business = m_headers
                m_col_idx = {h: col_map_m[h] for h in m_business if h in col_map_m}
                for h in schema_cols:
                    if h not in m_col_idx:
                        next_c = max(m_col_idx.values(), default=0) + 1
                        m_col_idx[h] = next_c
                        ws_master.cell(row=1, column=next_c, value=h)
                ws_master_meta = _get_or_create_meta_sheet(wb_master, sys_cols)
                m_data_count = max(0, (ws_master.max_row or 1) - 1)
                m_meta_count = max(0, (ws_master_meta.max_row or 1) - 1)
                m_sys_in_sheet = [h for h in m_headers if h in SYSTEM_COLUMNS]
                if m_data_count > 0 and m_meta_count == 0 and m_sys_in_sheet:
                    # Migrate master: copy system from Sheet1 to __link_meta
                    col_map_m = {h: i + 1 for i, h in enumerate(m_headers)}
                    for r in range(2, 2 + m_data_count):
                        for si, sc in enumerate(sys_cols, start=1):
                            cidx = col_map_m.get(sc)
                            if cidx is not None:
                                ws_master_meta.cell(row=r, column=si, value=ws_master.cell(row=r, column=cidx).value)
                    # Remove system columns from master Sheet1
                    cols_to_del = [i + 1 for i, h in enumerate(m_headers) if h in SYSTEM_COLUMNS]
                    for c in reversed(cols_to_del):
                        ws_master.delete_cols(c, 1)
                    m_col_idx = {h: i + 1 for i, h in enumerate(m_business)}
                elif m_data_count > m_meta_count:
                    for r in range(2 + m_meta_count, 2 + m_data_count):
                        ws_master_meta.cell(row=r, column=1, value=str(uuid.uuid4()))
                        ws_master_meta.cell(row=r, column=2, value="")
                        ws_master_meta.cell(row=r, column=3, value="")
                        if len(sys_cols) > 3:
                            ws_master_meta.cell(row=r, column=4, value="")
                        if len(sys_cols) > 4:
                            ws_master_meta.cell(row=r, column=5, value="")
                m_meta_max = ws_master_meta.max_row or 1
                uuid_to_row = {}
                for r in range(2, m_meta_max + 1):
                    uid = ws_master_meta.cell(row=r, column=1).value
                    src = ws_master_meta.cell(row=r, column=2).value if len(sys_cols) > 1 else None
                    if uid:
                        uuid_to_row[str(uid).strip()] = (r, str(src or "").strip())

                rows_inserted = 0
                rows_updated = 0
                appended_rows = []

                for row_idx, row in enumerate(data_rows):
                    if not row or all(c is None or str(c).strip() == "" for c in row):
                        continue
                    meta = meta_rows[row_idx] if row_idx < len(meta_rows) else {}
                    ru = (meta.get(row_uuid_col) or get_val(row, row_uuid_col))
                    if not ru:
                        ru = str(uuid.uuid4())
                    ru = str(ru).strip()
                    src_child = meta.get(source_child_col) or get_val(row, source_child_col) or child_file_id
                    src_user = meta.get(source_user_col) or get_val(row, source_user_col) or (child_owner or "")
                    child_row_1based = header_row_idx + 1 + row_idx

                    if ru in uuid_to_row:
                        if sync_mode == "append":
                            continue
                        mr, existing_src = uuid_to_row[ru]
                        if existing_src and existing_src != child_file_id:
                            continue
                        for h in schema_cols:
                            mc = m_col_idx.get(h)
                            if mc is not None:
                                val = get_val(row, h)
                                ws_master.cell(row=mr, column=mc, value=val)
                        _write_meta_row(ws_master_meta, mr, {row_uuid_col: ru, source_child_col: src_child, source_user_col: src_user, last_synced_col: datetime.now().strftime("%Y-%m-%d %H:%M"), row_locked_at_col: ""}, sys_cols)
                        rows_updated += 1
                    else:
                        next_row = ws_master.max_row + 1
                        if next_row == 2:
                            for h in schema_cols:
                                m_col_idx.setdefault(h, len(m_col_idx) + 1)
                                ws_master.cell(row=1, column=m_col_idx[h], value=h)
                        for h in schema_cols:
                            mc = m_col_idx.get(h)
                            if mc is None:
                                mc = max(m_col_idx.values(), default=0) + 1
                                m_col_idx[h] = mc
                                ws_master.cell(row=1, column=mc, value=h)
                            ws_master.cell(row=next_row, column=mc, value=get_val(row, h))
                        _write_meta_row(ws_master_meta, next_row, {row_uuid_col: ru, source_child_col: src_child, source_user_col: src_user, last_synced_col: datetime.now().strftime("%Y-%m-%d %H:%M"), row_locked_at_col: ""}, sys_cols)
                        uuid_to_row[ru] = (next_row, child_file_id)
                        rows_inserted += 1
                        row_hash = _row_content_hash(row, headers_ordered, row_locked_at_col)
                        appended_rows.append((child_row_1based, ru, row_hash))

                wb_master.save(master_path)
                wb_master.close()

                if appended_rows:
                    try:
                        wb_child = load_workbook(child_path, read_only=False, data_only=False)
                        ws_meta_w = wb_child[META_SHEET]
                        locked_at_val = datetime.now().strftime("%Y-%m-%d %H:%M")
                        locked_col_idx = sys_cols.index(row_locked_at_col) + 1 if row_locked_at_col in sys_cols else 5
                        for child_row_1based, ru, row_hash in appended_rows:
                            meta_row = 1 + (child_row_1based - header_row_idx - 1)
                            if meta_row >= 2:
                                ws_meta_w.cell(row=meta_row, column=locked_col_idx, value=locked_at_val)
                        wb_child.save(child_path)
                        wb_child.close()
                        db = get_db()
                        for _cidx, ru, row_hash in appended_rows:
                            db.execute(
                                "INSERT OR REPLACE INTO synced_row_locks (child_file_id, row_uuid, row_content_hash, locked_at) VALUES (?,?,?,?)",
                                (child_file_id, ru, row_hash, locked_at_val)
                            )
                        db.commit()
                        db.close()
                    except Exception as e:
                        logging.warning("excel_links set row_locked_at/synced_row_locks failed: %s", e)

                version_no = None
                version_created = False
                if rows_inserted > 0 or rows_updated > 0:
                    try:
                        from modules.files import create_version
                        os.makedirs(VERSIONS_DIR, exist_ok=True)
                        version_no = create_version(
                            master_file_id, master_path,
                            child_owner or "system",
                            "child_sync",
                            VERSIONS_DIR,
                            version_type="child_sync",
                            notes=f"Child {child_file_id}: +{rows_inserted} -{rows_updated}"
                        )
                        version_created = True
                    except Exception as verr:
                        logging.warning("excel_links: create_version after sync failed: %s", verr)

                _log_sync(master_file_id, child_file_id, "success", rows_inserted, rows_updated, 0, None, version_no)
                if not stored_schema_hash and schema_cols:
                    try:
                        new_hash = _compute_schema_hash(schema_cols)
                        db = get_db()
                        db.execute("UPDATE excel_link_schema SET schema_hash=? WHERE master_file_id=?", (new_hash, master_file_id))
                        db.commit()
                        db.close()
                    except Exception:
                        pass
                return {
                    "ok": True,
                    "rows_inserted": rows_inserted,
                    "rows_updated": rows_updated,
                    "rows_deleted": 0,
                    "version_created": version_created,
                    "version_no": version_no,
                }
            except Exception as e:
                logging.exception("excel_links sync_child_to_master failed: %s", e)
                _log_sync(master_file_id, child_file_id, "error", 0, 0, 0, str(e), None)
                return {"ok": False, "error": str(e), "version_created": False}
    finally:
        _release_file_lock(fd, lock_path)


def _log_sync(master_file_id, child_file_id, status, rows_inserted, rows_updated, rows_deleted, error_message, version_no=None):
    db = get_db()
    try:
        db.execute("""
            INSERT INTO excel_link_sync_log
            (master_file_id, child_file_id, status, rows_inserted, rows_updated, rows_deleted, error_message, synced_at, version_no)
            VALUES (?,?,?,?,?,?,?,?,?)
        """, (master_file_id, child_file_id, status, rows_inserted, rows_updated, rows_deleted, error_message or "", datetime.now().strftime("%Y-%m-%d %H:%M"), version_no))
        db.commit()
    except Exception as e:
        try:
            db.execute("""
                INSERT INTO excel_link_sync_log
                (master_file_id, child_file_id, status, rows_inserted, rows_updated, rows_deleted, error_message, synced_at)
                VALUES (?,?,?,?,?,?,?,?)
            """, (master_file_id, child_file_id, status, rows_inserted, rows_updated, rows_deleted, error_message or "", datetime.now().strftime("%Y-%m-%d %H:%M")))
            db.commit()
        except Exception:
            db.rollback()
        logging.warning("excel_links _log_sync version_no insert failed (fallback): %s", e)
    finally:
        db.close()


# ---- API helpers ----

def create_link(master_file_id, child_file_id, child_owner_user_id=None, child_owner_email=None, sheet_name="Sheet1", columns_json=None, sync_mode="append", header_row_index=1):
    """Create or re-activate file_links + excel_link_schema. Uses INSERT OR REPLACE to handle existing (e.g. paused) links."""
    db = get_db()
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        db.execute("""
            INSERT OR REPLACE INTO file_links (master_file_id, child_file_id, child_owner_user_id, child_owner_email, link_status, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?)
        """, (master_file_id, child_file_id, child_owner_user_id, child_owner_email, "active", now, now))
        sync_mode = (sync_mode or "append").strip().lower()
        if sync_mode not in ("append", "upsert"):
            sync_mode = "append"
        hdr_idx = header_row_index if header_row_index in (1, 2, 3) else 1
        business_cols = _get_schema_columns(columns_json)
        schema_hash_val = _compute_schema_hash(business_cols)
        db.execute("""
            INSERT OR REPLACE INTO excel_link_schema (master_file_id, sheet_name, columns_json, row_uuid_column_name, source_child_column_name, source_user_column_name, last_synced_at_column_name, row_locked_at_column_name, sync_mode, schema_hash, header_row_index, created_at, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (master_file_id, sheet_name, columns_json or "[]", "_row_uuid", "_source_child_file_id", "_source_user_id", "_last_synced_at", "_row_locked_at", sync_mode, schema_hash_val, hdr_idx, datetime.now().strftime("%Y-%m-%d %H:%M"), datetime.now().strftime("%Y-%m-%d %H:%M")))
        db.commit()
        return True
    except Exception as e:
        db.rollback()
        raise e
    finally:
        db.close()


def list_links(master_file_id):
    db = get_db()
    rows = db.execute("""
        SELECT fl.*,
               (SELECT status FROM excel_link_sync_log WHERE master_file_id=fl.master_file_id AND child_file_id=fl.child_file_id ORDER BY synced_at DESC LIMIT 1) as last_sync_status,
               (SELECT synced_at FROM excel_link_sync_log WHERE master_file_id=fl.master_file_id AND child_file_id=fl.child_file_id ORDER BY synced_at DESC LIMIT 1) as last_synced_at,
               (SELECT error_message FROM excel_link_sync_log WHERE master_file_id=fl.master_file_id AND child_file_id=fl.child_file_id ORDER BY synced_at DESC LIMIT 1) as last_error,
               (SELECT version_no FROM excel_link_sync_log WHERE master_file_id=fl.master_file_id AND child_file_id=fl.child_file_id AND status='success' ORDER BY synced_at DESC LIMIT 1) as last_version_no
        FROM file_links fl
        WHERE fl.master_file_id = ? AND fl.link_status = 'active'
    """, (master_file_id,)).fetchall()
    db.close()
    return [dict(r) for r in rows]


def unlink(master_file_id, child_file_id):
    """Set link_status to paused (V1: stop syncing, do not delete rows)."""
    db = get_db()
    db.execute("UPDATE file_links SET link_status='paused', updated_at=? WHERE master_file_id=? AND child_file_id=?", (datetime.now().strftime("%Y-%m-%d %H:%M"), master_file_id, child_file_id))
    db.commit()
    db.close()


def get_schema(master_file_id):
    db = get_db()
    row = db.execute("SELECT * FROM excel_link_schema WHERE master_file_id=?", (master_file_id,)).fetchone()
    db.close()
    return dict(row) if row else None


def get_sync_logs(master_file_id, child_file_id=None, limit=20):
    db = get_db()
    if child_file_id:
        rows = db.execute("""
            SELECT * FROM excel_link_sync_log WHERE master_file_id=? AND child_file_id=? ORDER BY synced_at DESC LIMIT ?
        """, (master_file_id, child_file_id, limit)).fetchall()
    else:
        rows = db.execute("""
            SELECT * FROM excel_link_sync_log WHERE master_file_id=? ORDER BY synced_at DESC LIMIT ?
        """, (master_file_id, limit)).fetchall()
    db.close()
    return [dict(r) for r in rows]


def is_child_linked(child_file_id):
    return get_link_by_child(child_file_id) is not None


def get_child_master(child_file_id):
    link = get_link_by_child(child_file_id)
    return link["master_file_id"] if link else None
